import os
import shutil
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime

# === Constants ===
EXCEL_TEMPLATE_PATH = "template/sales_template.xlsx"
INPUT_CSV_PATH = "data/data.csv"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === GUI Setup ===
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("Excel Report Generator")
app.geometry("600x420")

log_box = ctk.CTkTextbox(app, width=550, height=200)
log_box.pack(pady=20)

def log(msg):
    timestamp = datetime.now().strftime("[%H:%M:%S]")
    log_box.insert("end", f"{timestamp} {msg}\n")
    log_box.see("end")

# === Functions ===
def select_csv():
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if not file_path:
        log("No file selected.")
        return

    try:
        shutil.copyfile(file_path, INPUT_CSV_PATH)
        log(f"✅ Replaced input.csv with: {os.path.basename(file_path)}")
    except Exception as e:
        log(f"❌ Error copying file: {e}")

def generate_excel_report():
    if not os.path.exists(EXCEL_TEMPLATE_PATH):
        log("❌ Excel template not found.")
        return

    try:
        df = pd.read_csv(INPUT_CSV_PATH)
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
        ws = wb.active

        start_row = 2  # Assuming headers are in row 1
        for i, row in df.iterrows():
            for j, val in enumerate(row, start=1):
                ws.cell(row=start_row + i, column=j, value=val)

        output_path = os.path.join(OUTPUT_DIR, f"Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(output_path)
        log(f"📄 Excel report generated: {os.path.basename(output_path)}")
    except Exception as e:
        log(f"❌ Failed to generate Excel report: {e}")

# === Buttons ===
btn_csv = ctk.CTkButton(app, text="Select CSV File", command=select_csv)
btn_csv.pack(pady=10)

btn_generate = ctk.CTkButton(app, text="Generate Excel Report", command=generate_excel_report)
btn_generate.pack(pady=10)

app.mainloop()
